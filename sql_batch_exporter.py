# Importiere benötigte Bibliotheken
import os
import glob
from datetime import datetime  # Zum Verwalten von Zeitstempeln
from dotenv import load_dotenv
from sqlalchemy import create_engine
import urllib
import pandas as pd
from openpyxl.utils import get_column_letter
import logging
############################################################################################################
# Lade die .env - Datei
load_dotenv()  
############################################################################################################
# Dateipfade aus der .env-Datei
# Input
input_path=os.getenv("INPUT_PATH")
# Output
output_path=os.getenv("OUTPUT_PATH")
# Logfile-Pfad
logfile_path=os.getenv("LOGGER_PATH")
############################################################################################################
# LOGGER SETUP
log_dateiname = f"log_{datetime.now().strftime('%Y-%m-%d')}.log"
log_pfad = os.path.join(logfile_path, log_dateiname)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_pfad, encoding="utf-8"),  # → in Datei schreiben
        logging.StreamHandler()                            # → weiterhin in Konsole ausgeben
    ]
)

logger = logging.getLogger(__name__)
############################################################################################################


def verbinde_mit_sql_datenbank():
    """
    Verbindung mit der SQL Datenbank. Wahlweise mit Trusted Connection oder per Übergabe von Benutzernname und Passwort. 
    Secrets sind in der .env-Datei hinterlegt.
    """
    db_server = os.getenv("DB_SERVER")
    db_database = os.getenv("DATABASE")
    trusted_connection = os.getenv("TRUSTED_CONNECTION", "no").strip().lower()
    logger.info(f"Verbinde mit Server: {db_server} - Datenbank: {db_database}")
    if trusted_connection == "yes":
        params = urllib.parse.quote_plus(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={db_server};"
            f"DATABASE={db_database};"
            f"Trusted_Connection=yes;"
        )
        logger.info("🔐 Authentifizierung via Windows Authentication")
    else:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        params = urllib.parse.quote_plus(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={db_server};"
            f"DATABASE={db_database};"
            f"UID={db_user};"
            f"PWD={db_password};"
        )
        logger.info("🔐 Authentifizierung via Benutzername & Passwort")

    try:
        engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
        logger.info(f"✅ Engine für {db_database} erstellt.")
        return engine  # Engine zurückgeben, kein direktes conn-Objekt
    except Exception as e:
        logger.error(f"❌ Fehler bei Verbindung zur SQL-Datenbank: {db_database} Fehler: {e}")
        raise
      
      
def lade_sql_dateien(input_path):
    """
    Lese mehrere SQL-Dateien im Ordner ein und gib den Inhalt als Dictionary zurück.
    Key = Dateiname, Value = SQL-Inhalt 
    """
    sql_queries = {}
    
    # Alle .sql Dateien im Ordner finden
    sql_dateien = glob.glob(os.path.join(input_path, "*.sql"))
    # Encoding aus der .env Datei
    encoding=os.getenv("ENCODING")
    
    # Prüfe ob die richtige Dateiendung .sql im Ordner vorhanden ist
    if not sql_dateien:
        logger.error(f"❌ Keine SQL-Dateien gefunden in: {input_path}")
        raise FileNotFoundError(f"❌ Keine SQL-Dateien gefunden in: {input_path}")
    
    for path in sql_dateien:
        try:
            logger.info("Versuche SQL-Dateien einzulesen...")
            with open (path, "r", encoding=encoding) as file:
                dateiname = os.path.basename(path) 
                sql_queries[dateiname]=file.read()
            logger.info(f"✅ SQL Datei erfolgreich eingelesen: {dateiname}")
        except Exception as e:
            logger.error(f"❌ Fehler beim Einlesen der SQL Dateien: {path} : {e}")
            continue

    # Rückgabe als Dictionary    
    return sql_queries


def sql_dataframe_erstellen(sql_queries, conn):  
    """
    Erstelle einen DataFrame pro SQL-Datei.
    Key = Dateiname, Value = DataFrame
    """
    dataframes = {}
    
    for dateiname, query in sql_queries.items():
        try:
            logger.info("Versuche DataFrame(s) zu erstellen...")
            # erstelle DataFrame(s) aus den SQL-Datieien
            df = pd.read_sql(query, conn)
            dataframes[dateiname] = df
            logger.info(f"✅ DataFrame(s) {dateiname} erfolgreich erstellt.")
        except Exception as e:
            logger.error(f"❌ Fehler beim erstellen des DataFrames: {dateiname} - Fehler: {e}")
            continue
    # Rückgabe als Dictionary      
    return dataframes 


def export_to_excel(dataframes, output_path):
    """
    Exportiert alle DataFrames als einzelne Excel-Dateien.

    Parameter:
        dataframes (dict): Dictionary mit Dateiname -> DataFrame
        output_path (str): Zielordner
    """
    for dateiname, df in dataframes.items():  # ueber Dictionary iterieren
        try:
            # Dateiname: "abfrage1.sql" -> "abfrage1.xlsx"
            excel_dateiname = os.path.splitext(dateiname)[0] + ".xlsx"
            full_output_path = os.path.join(output_path, excel_dateiname)  # ✅ vollständiger Pfad
            logger.info(f"⚙️ Starte Export: {dateiname}")
            
            # Zeitstempel in letzte Spalte einfügen
            df['Export_Zeitstempel'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  
            
            with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Daten')
                # Tabellenblattname
                worksheet = writer.sheets['Daten']
                # Erste Zeile fixiert
                worksheet.freeze_panes = worksheet['A2']
                # Spaltenbreite
                for i, column in enumerate(df.columns, 1):  
                    max_length = max(
                        df[column].astype(str).map(len).max(),
                        len(column)
                    )
                    adjusted_width = min(max_length + 5, 30)
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = adjusted_width  
                # Setze Auto-Filter in die Datei    
                worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"  
                    
            logger.info(f"✅ Abfragen Erfolgreich exportiert: {full_output_path}")
            
        except PermissionError:
            logger.error(f"❌ Fehler: '{excel_dateiname}' ist noch geöffnet. Bitte die Datei schließen.")
            continue
        except Exception as e:
            logger.error(f"❌ Fehler beim Export von '{dateiname}': {e}")
            continue


def export_to_csv(dataframes, output_path):
    """
    Exportiert alle DataFrames als einzelne CSV-Dateien.

    Parameter:
        dataframes (dict): Dictionary mit Dateiname -> DataFrame
        output_path (str): Zielordner
    """
    for dateiname, df in dataframes.items():
        try:
            # Dateiname: "abfrage1.sql" -> "abfrage1.csv"
            csv_dateiname = os.path.splitext(dateiname)[0] + ".csv"
            full_output_path = os.path.join(output_path, csv_dateiname)
            logger.info(f"⚙️ Starte Export: {dateiname}")
            
            # Zeitstempel in letzte Spalte einfügen
            df['Export_Zeitstempel'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Direkt als CSV speichern
            df.to_csv(full_output_path, index=False, sep=';', encoding='utf-8-sig')
            #                                    Semikolon       BOM für Excel-Kompatibilität
                    
            logger.info(f"✅ Abfragen Erfolgreich exportiert: {full_output_path}")
            
        except PermissionError:
            logger.error(f"❌ Fehler: '{csv_dateiname}' ist noch geöffnet. Bitte Datei schließen.")      
            continue
        except Exception as e:
            logger.error(f"❌ Fehler beim Export von '{dateiname}': {e}")
            continue

def close_connection(conn):
    """
    Versuche die Verbindung zur Datenbank zu trennen.
    """
    try:
        conn.close()
        logger.info(f"✅ Verbindung zur Datenbank wurde geschlossen.")
        logger.info(f"✅ PROZESS BEENDET.")
    except Exception as e:
        logger.error(f"❌ Verbindung zur Datenbank konnte nicht geschlossen werden.")
        raise     

# Aufruf der Funktionen
engine = verbinde_mit_sql_datenbank()
queries = lade_sql_dateien(input_path)

# Connection aus der Engine holen
try:
    conn = engine.connect()
    logger.info("✅ Connection aus Engine erfolgreich erstellt.")
except Exception as e:
    logger.error(f"❌ Fehler beim Erstellen der Connection: {e}")
    raise

dataframes = sql_dataframe_erstellen(queries, conn)
export_to_excel(dataframes, output_path)

#-> Bei Bedarf alternativer Export als CSV-Datei
#export_to_csv(dataframes, output_path)

# Schließe die Verbindung zur Datenbank auf dem Server
close_connection(conn)
